"""Import d'un classeur Excel de dépenses.

Deux classeurs sont lus par le même code :

- **l'export de la plateforme** (13 colonnes, en-tête en première ligne) ;
- **le classeur historique du client** (« BASE DE DONNEES ACTIONS » : un
  titre fusionné, une note, puis l'en-tête en septième ligne et 9 colonnes —
  N°ORDRE, DATE, TEAM, OWNER, LIBELLE DES TRANSACTIONS, DEPENSES, MONTANT
  JUSTIFIER, ECART, PIECES JUSTIFICATIVES). Ce fichier est mono-pays : le
  pays vient alors de la requête, pas du classeur.

Tout ce qui entre par ici arrive en brouillon, sans montant justifié : le
classeur déclare, le siège constate. MONTANT JUSTIFIER et ECART sont donc
ignorés ; la mention de la pièce (« Reçu », « Reçu(justif incomplet) ») est
conservée en remarque de la ligne, comme une information — pas comme une
preuve. Chaque ligne est validée séparément et signalée par son numéro de
ligne dans le classeur ; rien n'est écrit tant qu'une seule ligne est en
erreur.
"""

import logging
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db import IntegrityError, transaction
from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import load_workbook

from accounts.permissions import get_access
from budget.aggregates import convert
from core.journal import tracer
from core.models import Country, Manager, Team
from expenses.models import AuditLog, Dossier, Expense
from expenses.workflow import Status

from .scope import fuseau_de

logger = logging.getLogger(__name__)

# openpyxl lit le XML du classeur avec ``defusedxml`` dès que le paquet est
# installé, ce qui neutralise les entités externes et les « billion laughs »
# qu'un classeur forgé pourrait contenir. L'import est protégé : le code
# fonctionne sans, mais le dit, pour que l'absence se voie dans les journaux
# plutôt qu'au premier classeur malveillant.
try:
    import defusedxml  # noqa: F401

    DEFUSEDXML_DISPONIBLE = True
except ImportError:  # pragma: no cover - dépend de l'environnement
    DEFUSEDXML_DISPONIBLE = False
    logger.warning(
        "defusedxml n'est pas installé : les classeurs importés sont lus "
        "sans protection contre les entités XML."
    )

#: Colonnes sans lesquelles une ligne ne se déclare pas. Ce sont celles du
#: classeur historique : l'export de la plateforme les porte aussi.
COLONNES_OBLIGATOIRES = [
    "N°ORDRE", "DATE", "TEAM", "OWNER", "LIBELLE DES TRANSACTIONS", "DEPENSES",
]

#: Colonnes lues quand elles existent. PAYS manque au classeur historique
#: (mono-pays) ; la devise d'origine n'y figure pas non plus. MONTANT
#: JUSTIFIER, ECART et STATUT ne sont pas lus du tout : le siège constate.
COLONNES_FACULTATIVES = [
    "PAYS", "DEVISE D'ORIGINE", "MONTANT D'ORIGINE", "PIECES JUSTIFICATIVES",
]

#: Deux colonnes dont la présence signe la ligne d'en-tête : elles figurent
#: dans les deux formats et dans aucun titre ni aucune note.
MARQUEURS_D_ENTETE = ("N°ORDRE", "DEPENSES")

#: Lignes parcourues à la recherche de l'en-tête. Le classeur historique le
#: place en septième ligne ; au-delà de quinze, ce n'est plus un titre mais
#: un autre fichier.
LIGNES_D_ENTETE_MAX = 15

#: Nombre maximal de lignes par classeur. Au-delà, ce n'est plus une saisie
#: mais une reprise de données, qui ne doit pas passer par une requête web.
LIGNES_MAX = int(getattr(settings, "IMPORT_MAX_ROWS", 5000))

#: Taille des lots d'insertion : assez grand pour limiter les allers-retours,
#: assez petit pour que la requête reste raisonnable.
TAILLE_LOT = 500

#: Longueurs des champs texte, reprises du modèle : une valeur trop longue
#: doit être refusée ligne par ligne, pas par une erreur de base au moment
#: de l'écriture, qui perdrait tout le classeur.
LONGUEURS = {
    "N°ORDRE": Dossier._meta.get_field("number").max_length,
    "LIBELLE DES TRANSACTIONS": Expense._meta.get_field("title").max_length,
    "DEVISE D'ORIGINE": Expense._meta.get_field("original_currency").max_length,
    "TEAM": Team._meta.get_field("name").max_length,
    "OWNER": Manager._meta.get_field("name").max_length,
}

#: Chiffres avant la virgule autorisés par ``DecimalField(16, 2)``.
CHIFFRES_ENTIERS_MAX = (
    Expense._meta.get_field("amount").max_digits
    - Expense._meta.get_field("amount").decimal_places
)
CENTS = Decimal("0.01")

#: Formats de date acceptés en texte : l'export écrit l'heure, le classeur
#: historique n'en a pas.
FORMATS_DE_DATE = ("%d/%m/%Y %H:%M", "%d/%m/%Y")


def _texte(value):
    return "" if value is None else str(value).strip()


def _texte_borne(row, entete):
    valeur = _texte(row[entete])
    if len(valeur) > LONGUEURS[entete]:
        raise ValueError(
            _("%(entete)s trop long (%(taille)s caractères, maximum %(max)s).")
            % {"entete": entete, "taille": len(valeur), "max": LONGUEURS[entete]}
        )
    return valeur


def _numero_d_ordre(row):
    """Le N°ORDRE en texte, tel qu'un humain l'écrirait.

    Le classeur historique le porte en nombre entier ; une cellule numérique
    relue en flottant donnerait « 12.0 », qui ne rejoindrait jamais le
    dossier « 12 » créé à la main.
    """
    valeur = row["N°ORDRE"]
    if isinstance(valeur, float) and valeur.is_integer():
        valeur = int(valeur)
    if isinstance(valeur, int) and not isinstance(valeur, bool):
        valeur = str(valeur)
    numero = _texte(valeur)
    if len(numero) > LONGUEURS["N°ORDRE"]:
        raise ValueError(
            _("%(entete)s trop long (%(taille)s caractères, maximum %(max)s).")
            % {"entete": "N°ORDRE", "taille": len(numero), "max": LONGUEURS["N°ORDRE"]}
        )
    if not numero:
        raise ValueError(_("N°ORDRE obligatoire."))
    return numero


def _montant(value, entete):
    """Montant positif et fini, ou ``None`` si la cellule est vide.

    ``Decimal`` accepte « NaN » et « Infinity » sans broncher ; la base, non.
    Une telle valeur — ou un montant de plus de quatorze chiffres — doit
    devenir une erreur de ligne, jamais une erreur 500 à l'écriture.
    """
    if value in (None, ""):
        return None
    try:
        montant = Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, AttributeError, ValueError):
        raise ValueError(_("%(entete)s illisible : « %(value)s »") % {"entete": entete, "value": value})
    if not montant.is_finite() or montant < 0:
        raise ValueError(_("%(entete)s illisible : « %(value)s »") % {"entete": entete, "value": value})
    montant = montant.quantize(CENTS)
    if montant.adjusted() + 1 > CHIFFRES_ENTIERS_MAX:
        raise ValueError(
            _("%(entete)s trop grand : « %(value)s » (maximum %(max)s chiffres avant la virgule).")
            % {"entete": entete, "value": value, "max": CHIFFRES_ENTIERS_MAX}
        )
    return montant


def _date(value, fuseau):
    """Date de la ligne, avec ou sans heure, dans le fuseau du pays.

    Le classeur historique ne porte que le jour : la dépense est alors datée
    de minuit. Le classeur est écrit à l'heure du pays — c'est celle qu'on
    lit sur les pièces et celle de l'export — et se relit dans ce même
    fuseau : la faire passer par celui du serveur décalait une ligne du
    1er janvier à 01:00 dans l'exercice précédent.
    """
    if isinstance(value, datetime):
        return value.replace(tzinfo=fuseau) if timezone.is_naive(value) else value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=fuseau)
    texte = _texte(value)
    for format_ in FORMATS_DE_DATE:
        try:
            return datetime.strptime(texte, format_).replace(tzinfo=fuseau)
        except ValueError:
            continue
    raise ValueError(_("Date illisible : « %(value)s »") % {"value": value})


def _ouvrir(uploaded):
    taille = getattr(uploaded, "size", None)
    if taille is not None and taille > settings.MAX_PROOF_SIZE:
        limite = settings.MAX_PROOF_SIZE // (1024 * 1024)
        raise ValueError(_("Classeur trop volumineux (maximum %(limite)s Mo).") % {"limite": limite})
    # Un xlsx est une archive : quelques Mo compressés peuvent en cacher
    # des Go de chaînes partagées, qu'openpyxl chargerait en mémoire.
    limite = 5 * settings.MAX_PROOF_SIZE
    try:
        with zipfile.ZipFile(uploaded) as archive:
            decompresse = sum(info.file_size for info in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise ValueError(_("Classeur illisible : ce n'est pas un fichier xlsx.")) from exc
    uploaded.seek(0)
    if decompresse > limite:
        raise ValueError(
            _("Classeur trop volumineux une fois décompressé (maximum %(limite)s Mo).")
            % {"limite": limite // (1024 * 1024)}
        )
    try:
        return load_workbook(uploaded, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(_("Classeur illisible : %(erreur)s") % {"erreur": exc}) from exc


def _entete(value):
    """Libellé de colonne normalisé : espaces repliés, casse ignorée."""
    return re.sub(r"\s+", " ", _texte(value)).upper()


def _trouver_l_entete(rows):
    """Cherche la ligne d'en-tête dans les premières lignes du classeur.

    Renvoie ``(numéro de ligne, positions des colonnes)``. L'export met
    l'en-tête en première ligne ; le classeur historique l'a en septième,
    sous un titre fusionné et une note. Reconnaître l'en-tête à son contenu
    plutôt qu'à sa place permet de lire les deux — et un classeur remanié.
    """
    for numero, row in enumerate(rows, start=1):
        if numero > LIGNES_D_ENTETE_MAX:
            break
        entetes = [_entete(value) for value in row]
        if all(marqueur in entetes for marqueur in MARQUEURS_D_ENTETE):
            positions = {
                colonne: entetes.index(colonne)
                for colonne in COLONNES_OBLIGATOIRES + COLONNES_FACULTATIVES
                if colonne in entetes
            }
            manquantes = [c for c in COLONNES_OBLIGATOIRES if c not in positions]
            if manquantes:
                raise ValueError(
                    _("En-têtes manquants : %(colonnes)s") % {"colonnes": ", ".join(manquantes)}
                )
            return numero, positions
    raise ValueError(
        _(
            "Ligne d'en-tête introuvable : le classeur doit porter les colonnes "
            "%(colonnes)s dans ses %(max)s premières lignes."
        ) % {"colonnes": ", ".join(COLONNES_OBLIGATOIRES), "max": LIGNES_D_ENTETE_MAX}
    )


def _charger_lignes(uploaded):
    """Lit le classeur et renvoie ``(lignes, colonnes présentes)``.

    Chaque ligne est ``(numéro dans le classeur, {colonne: valeur})`` : le
    numéro est celui qu'Excel affiche, pour que l'erreur signalée se
    retrouve dans le fichier.
    """
    workbook = _ouvrir(uploaded)
    sheet = (
        workbook["BASE DE DONNEES ACTIONS"]
        if "BASE DE DONNEES ACTIONS" in workbook.sheetnames
        else workbook.active
    )
    rows = sheet.iter_rows(values_only=True)
    ligne_d_entete, positions = _trouver_l_entete(rows)

    def cellule(row, colonne):
        position = positions.get(colonne)
        return row[position] if position is not None and position < len(row) else None

    lignes = []
    for line_number, row in enumerate(rows, start=ligne_d_entete + 1):
        if not any(value not in (None, "") for value in row):
            continue
        if (
            _texte(cellule(row, "N°ORDRE")) == ""
            and _texte(cellule(row, "LIBELLE DES TRANSACTIONS")).upper() == "TOTAL"
        ):
            continue
        if len(lignes) >= LIGNES_MAX:
            raise ValueError(
                _("Le classeur dépasse %(max)s lignes : scindez-le.") % {"max": LIGNES_MAX}
            )
        lignes.append(
            (
                line_number,
                {
                    colonne: cellule(row, colonne)
                    for colonne in COLONNES_OBLIGATOIRES + COLONNES_FACULTATIVES
                },
            )
        )
    return lignes, set(positions)


def _erreur(ligne, motif):
    return {"ligne": ligne, "motif": motif}


def _managers_par_pays():
    """Managers actifs, indexés par (pays, nom).

    Un manager n'existe que rattaché à un pays : résoudre « Kodjo Mensah » par
    son seul nom rattachait la ligne au premier homonyme trouvé, fût-il chez
    le voisin.
    """
    index = {}
    for manager in Manager.objects.filter(is_active=True).prefetch_related("countries"):
        for country in manager.countries.all():
            index.setdefault((country.pk, manager.name.casefold()), manager)
    return index


def _devise_d_origine(row, country, date, amount):
    """Applique la règle « les deux ou aucun » du §5.3 et fige le taux.

    Renvoie ``(amount, original_currency, original_amount, original_rate)``.
    Quand la pièce est libellée dans une autre devise, c'est la conversion —
    au taux du jour de la dépense — qui pèse sur l'enveloppe : DEPENSES est
    alors recalculée plutôt que reprise du classeur, pour que la ligne
    importée obéisse à la même règle qu'une ligne saisie.
    """
    devise = _texte_borne(row, "DEVISE D'ORIGINE").upper()
    montant = _montant(row["MONTANT D'ORIGINE"], "MONTANT D'ORIGINE")
    if not devise and montant is None:
        if amount is None:
            raise ValueError(_("Montant de dépense obligatoire."))
        return amount, "", None, None
    if not devise or montant is None:
        raise ValueError(
            _("Indiquez à la fois la devise et le montant d'origine, ou aucun des deux.")
        )
    if devise == country.currency:
        return montant, "", None, None
    converti, taux = convert(montant, devise, country.currency, date.date())
    if converti is None:
        raise ValueError(
            _("Aucun taux connu pour convertir %(devise)s en %(cible)s au %(date)s.")
            % {"devise": devise, "cible": country.currency, "date": date.date().strftime("%d/%m/%Y")}
        )
    return converti, devise, montant, taux


def _note(row):
    """La mention de pièce du classeur, gardée comme information.

    « Reçu » ou « Reçu(justif incomplet) » dans le fichier historique ne
    prouve rien : la pièce elle-même n'est pas dans le classeur. La mention
    est conservée en remarque pour que le contrôleur sache qu'une pièce
    existait au moment de la saisie, et la réclame.
    """
    piece = _texte(row["PIECES JUSTIFICATIVES"])
    return f"Pièce : {piece}" if piece else ""


def _empreinte(number, jour, title, amount):
    """Ce qui fait qu'une ligne « existe déjà » : dossier, jour, libellé, montant.

    Le jour, pas l'instant : le classeur ne porte que la date, alors qu'une
    ligne saisie dans l'application porte l'heure. Comparer les instants
    faisait recréer, à chaque réimport d'un export, toute ligne saisie
    ailleurs qu'à minuit.
    """
    return (number, jour, title, amount)


def _lignes_en_base(dossier, cache):
    """Empreintes des lignes déjà présentes dans un dossier, lues une fois."""
    if dossier.pk not in cache:
        fuseau = fuseau_de(dossier.country)
        cache[dossier.pk] = {
            _empreinte(dossier.number, timezone.localtime(instant, fuseau).date(), title, amount)
            for instant, title, amount in dossier.expenses.values_list("date", "title", "amount")
        }
    return cache[dossier.pk]


def _resoudre_le_pays(row, avec_colonne_pays, pays_par_nom, pays_impose, access):
    """Le pays de la ligne : la colonne PAYS, à défaut celui de la requête."""
    pays_nom = _texte(row["PAYS"]) if avec_colonne_pays else ""
    if not pays_nom:
        if pays_impose is None:
            raise ValueError(_("Pays obligatoire : la colonne PAYS est vide."))
        return pays_impose
    country = pays_par_nom.get(pays_nom.casefold())
    if country is None:
        raise ValueError(_("Pays « %(pays)s » inconnu") % {"pays": pays_nom})
    if not access.has_global_scope and country.pk not in access.country_ids:
        raise ValueError(_("Pays « %(pays)s » hors périmètre") % {"pays": country.name})
    return country


def importer_depenses(uploaded, user, dry_run=False, country=None):
    """Valide tout le classeur, puis le crée atomiquement si demandé.

    ``country`` est le pays de l'import, déjà vérifié contre le périmètre
    par la vue. Il est obligatoire quand le classeur n'a pas de colonne
    PAYS — le cas du fichier historique — et sert de repli quand la cellule
    PAYS d'une ligne est vide.
    """
    try:
        lignes, colonnes = _charger_lignes(uploaded)
    except ValueError as exc:
        return _resultat(0, 0, [_erreur(1, str(exc))], dry_run)

    avec_colonne_pays = "PAYS" in colonnes
    if not avec_colonne_pays and country is None:
        return _resultat(
            0, 0,
            [_erreur(1, _(
                "Le classeur n'a pas de colonne PAYS : indiquez le pays "
                "de l'import (paramètre « country »)."
            ))],
            dry_run,
        )

    access = get_access(user)
    pays = {c.name.casefold(): c for c in Country.objects.all()}
    equipes = {(team.country_id, team.name.casefold()): team for team in Team.objects.all()}
    managers = _managers_par_pays()
    # Équipes et managers que le classeur nomme et que le pays ne connaît
    # pas encore : ils sont créés à l'écriture, une fois par nom. Le
    # classeur historique est la première source du référentiel — exiger
    # qu'il soit saisi à la main avant l'import rendrait l'import inutile.
    equipes_a_creer = {}
    managers_a_creer = {}
    erreurs = []
    valides = []
    dossiers_existants = {}
    # Lignes déjà en base ou déjà vues dans ce classeur : réimporter le même
    # fichier — ou le même classeur collé deux fois — ne doit rien créer.
    empreintes_vues = {}
    lignes_en_base = {}

    for numero_ligne, row in lignes:
        try:
            number = _numero_d_ordre(row)
            pays_ligne = _resoudre_le_pays(
                row, avec_colonne_pays, pays, country, access
            )

            date_ligne = _date(row["DATE"], fuseau_de(pays_ligne))
            amount = _montant(row["DEPENSES"], "DEPENSES")
            title = _texte_borne(row, "LIBELLE DES TRANSACTIONS") or "Dépense importée"
            amount, devise, montant_origine, taux = _devise_d_origine(
                row, pays_ligne, date_ligne, amount
            )
            team_name = _texte_borne(row, "TEAM")
            cle_equipe = (pays_ligne.pk, team_name.casefold()) if team_name else None
            team = equipes.get(cle_equipe) if team_name else None
            if team_name and team is None:
                equipes_a_creer.setdefault(cle_equipe, (pays_ligne, team_name))
            owner_name = _texte_borne(row, "OWNER")
            cle_manager = (pays_ligne.pk, owner_name.casefold()) if owner_name else None
            owner = managers.get(cle_manager) if owner_name else None
            if owner_name and owner is None:
                managers_a_creer.setdefault(cle_manager, (pays_ligne, owner_name))

            # Le N°ORDRE est unique par pays : le dossier se cherche dans le
            # pays de la ligne, jamais ailleurs.
            cle_dossier = (pays_ligne.pk, number)
            if cle_dossier not in dossiers_existants:
                dossiers_existants[cle_dossier] = Dossier.objects.filter(
                    country=pays_ligne, number=number
                ).first()
            dossier = dossiers_existants[cle_dossier]
            if dossier is not None and dossier.status != Status.DRAFT:
                raise ValueError(_("Le dossier « %(number)s » est déjà déclaré") % {"number": number})

            empreinte = _empreinte(cle_dossier, date_ligne.date(), title, amount)
            deja = empreintes_vues.get(empreinte)
            if deja is not None:
                raise ValueError(_("Ligne identique à la ligne %(ligne)s du classeur") % {"ligne": deja})
            if dossier is not None and _empreinte(
                number, date_ligne.date(), title, amount
            ) in _lignes_en_base(dossier, lignes_en_base):
                raise ValueError(
                    _(
                        "Ligne déjà présente dans le dossier « %(number)s » : "
                        "même date, même libellé, même montant"
                    ) % {"number": number}
                )
            empreintes_vues[empreinte] = numero_ligne

            valides.append({
                "ligne": numero_ligne,
                "cle_dossier": cle_dossier,
                "number": number,
                "country": pays_ligne,
                "team": team,
                "cle_equipe": cle_equipe,
                "owner": owner,
                "cle_manager": cle_manager,
                "date": date_ligne,
                "title": title,
                "amount": amount,
                "original_currency": devise,
                "original_amount": montant_origine,
                "original_rate": taux,
                "note": _note(row),
            })
        except ValueError as exc:
            erreurs.append(_erreur(numero_ligne, str(exc)))

    nouveaux_dossiers = len({
        ligne["cle_dossier"] for ligne in valides
        if dossiers_existants[ligne["cle_dossier"]] is None
    })
    resultat = _resultat(
        nouveaux_dossiers, len(valides), erreurs, dry_run,
        equipes_creees=len(equipes_a_creer), managers_crees=len(managers_a_creer),
    )
    if erreurs or dry_run:
        return resultat

    try:
        with transaction.atomic():
            _ecrire(valides, user, equipes, managers, equipes_a_creer, managers_a_creer,
                    dossiers_existants)
    except _LigneEnErreur as exc:
        # La transaction est défaite : rien n'a été écrit, comme pour une
        # erreur relevée à la validation.
        return _resultat(0, 0, [_erreur(exc.ligne, exc.motif)], dry_run)
    return resultat


class _LigneEnErreur(Exception):
    """Erreur relevée à l'écriture, rapportée à sa ligne du classeur."""

    def __init__(self, ligne, motif):
        super().__init__(motif)
        self.ligne = ligne
        self.motif = motif


def _creer_le_dossier(ligne, user):
    """Crée le dossier d'une ligne, ou signale qu'un autre import l'a fait.

    Le dossier était absent à la validation, mais deux imports du même
    classeur peuvent se croiser : le second heurtait la contrainte d'unicité
    (pays, N°ORDRE) et répondait 500, en ayant perdu tout le classeur.
    ``get_or_create`` absorbe la course ; si le dossier existe désormais, la
    ligne est refusée avec son numéro — ses doublons n'ont pas été vérifiés
    contre ce dossier-là — et l'import se relance.
    """
    motif = _(
        "Le dossier « %(number)s » vient d'être créé par un autre import : "
        "relancez l'import."
    ) % {"number": ligne["number"]}
    try:
        with transaction.atomic():
            dossier, cree = Dossier.objects.get_or_create(
                country=ligne["country"],
                number=ligne["number"],
                defaults={
                    "label": ligne["title"] or ligne["number"],
                    "team": ligne["team"],
                    "owner": ligne["owner"],
                    "date": ligne["date"].date(),
                    "status": Status.DRAFT,
                    "created_by": user.username,
                },
            )
    except IntegrityError:
        raise _LigneEnErreur(ligne["ligne"], motif)
    if not cree:
        raise _LigneEnErreur(ligne["ligne"], motif)
    return dossier


def _ecrire(valides, user, equipes, managers, equipes_a_creer, managers_a_creer,
            dossiers_existants):
    """Écrit référentiel, dossiers et lignes, dans la transaction de l'appelant."""
    # Le référentiel manquant d'abord : les lignes s'y rattachent.
    # ``create`` un par un, et non ``bulk_create`` : la création doit
    # passer par les signaux d'historisation (``ChangeLog``).
    for cle, (pays_equipe, nom) in equipes_a_creer.items():
        equipes[cle] = Team.objects.create(country=pays_equipe, name=nom)
    for cle, (pays_manager, nom) in managers_a_creer.items():
        manager = Manager.objects.create(name=nom)
        pays_manager.managers.add(manager)
        managers[cle] = manager

    dossiers = {}
    depenses = []
    for ligne in valides:
        if ligne["team"] is None and ligne["cle_equipe"] is not None:
            ligne["team"] = equipes[ligne["cle_equipe"]]
        if ligne["owner"] is None and ligne["cle_manager"] is not None:
            ligne["owner"] = managers[ligne["cle_manager"]]
        dossier = dossiers.get(ligne["cle_dossier"])
        if dossier is None:
            dossier = dossiers_existants[ligne["cle_dossier"]]
            if dossier is None:
                dossier = _creer_le_dossier(ligne, user)
            dossiers[ligne["cle_dossier"]] = dossier
        depenses.append(
            Expense(
                dossier=dossier,
                country=ligne["country"],
                team=ligne["team"],
                owner=ligne["owner"],
                date=ligne["date"],
                title=ligne["title"],
                amount=ligne["amount"],
                # Le classeur peut porter un MONTANT JUSTIFIER : il est
                # ignoré. Une preuve se constate au siège, elle ne
                # s'importe pas.
                justified_amount=Decimal("0.00"),
                original_currency=ligne["original_currency"],
                original_amount=ligne["original_amount"],
                original_rate=ligne["original_rate"],
                note=ligne["note"],
                status=Status.DRAFT,
                created_by=user.username,
            )
        )
    # Aucun signal n'écoute ``Expense`` : l'insertion par lots ne fait
    # perdre aucune trace, et évite une requête par ligne.
    Expense.objects.bulk_create(depenses, batch_size=TAILLE_LOT)


def _resultat(dossiers, lignes, erreurs, dry_run, *, equipes_creees=0, managers_crees=0):
    return {
        "dossiers_crees": dossiers,
        "lignes_creees": lignes,
        "equipes_creees": equipes_creees,
        "managers_crees": managers_crees,
        "erreurs": erreurs,
        "dry_run": dry_run,
    }


def audit_import(request, resultat, country=None):
    """Un import verse des lignes dans le système : il laisse une trace."""
    tracer(
        request,
        AuditLog.Action.IMPORTED,
        "ExpenseImport",
        famille="import",
        label="Import des dépenses Excel",
        # Le pays de l'import, quand il vient de la requête : le journal
        # d'un pays doit montrer ce qui y a été versé.
        country=country,
        **resultat,
    )
