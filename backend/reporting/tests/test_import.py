"""Import Excel des dépenses, dans le contrat produit par l'export."""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest import mock
from urllib.parse import urlencode

from django.db import IntegrityError, connection
from django.test import override_settings
from django.test.utils import CaptureQueriesContext
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from rest_framework import status

from accounts.models import Role
from accounts.tests.test_scoping import make_user
from budget.models import ExchangeRate
from core.models import ChangeLog, Manager, Team
from expenses.models import AuditLog, Dossier, Expense
from expenses.tests.base import ExpenseTestCase
from expenses.workflow import Status
from reporting import imports
from reporting.exports import EXPENSE_COLUMNS

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class ImportTests(ExpenseTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        # Un manager se résout dans son pays : celui du socle n'y était pas
        # rattaché.
        cls.togo.managers.add(cls.manager)

    def _classeur(self, lignes, entetes=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BASE DE DONNEES ACTIONS"
        entetes = entetes or [title for title, _ in EXPENSE_COLUMNS]
        sheet.append(entetes)
        for ligne in lignes:
            sheet.append([ligne.get(entete) for entete in entetes])
        contenu = BytesIO()
        workbook.save(contenu)
        contenu.seek(0)
        return contenu

    def _ligne(self, **overrides):
        ligne = {
            "N°ORDRE": "N-IMPORT-01",
            "DATE": f"15/03/{self.year} 12:30",
            "PAYS": "Togo",
            "TEAM": "Équipe Lomé",
            "OWNER": "Kodjo Mensah",
            "LIBELLE DES TRANSACTIONS": "Déplacement",
            "DEPENSES": 1200,
            "DEVISE D'ORIGINE": "",
            "MONTANT D'ORIGINE": "",
            "MONTANT JUSTIFIER": 800,
            "ECART": 400,
            "STATUT": "Brouillon",
            "PIECES JUSTIFICATIVES": "Facture",
        }
        ligne.update(overrides)
        return ligne

    def _importer(self, contenu, user=None, **query):
        self.login(user or self.doo)
        url = "/api/imports/expenses.xlsx"
        if query:
            url += "?" + urlencode(query)
        return self.client.post(
            url, {"file": ("depenses.xlsx", contenu, XLSX)}, format="multipart"
        )

    def test_un_classeur_exporte_se_reimporte(self):
        """Export et import partagent le même contrat de fichier, entre les
        mains des administrateurs — les seuls à manipuler des fichiers."""
        self.make_expense(amount="1200.00", justified_amount="800.00")
        self.login(self.doo)
        export = self.client.get("/api/exports/expenses.xlsx", {"year": self.year})
        Expense.objects.all().delete()
        Dossier.objects.all().delete()

        response = self._importer(BytesIO(export.content), user=self.doo)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["lignes_creees"], 1)
        self.assertFalse(response.data["erreurs"])
        expense = Expense.objects.get()
        self.assertEqual(str(expense.amount), "1200.00")
        # Le classeur portait 800 de justifié : l'import n'en tient pas compte.
        self.assertEqual(str(expense.justified_amount), "0.00")
        self.assertEqual(expense.status, Status.DRAFT)

    def test_seuls_les_administrateurs_importent(self):
        """Importer est réservé aux administrateurs : la direction financière
        constate, le pays déclare dans l'application — aucun d'eux ne
        manipule de fichier."""
        for user in (self.controller, self.rep_ivoire, self.owner):
            with self.subTest(role=user.profile.role):
                response = self._importer(self._classeur([self._ligne()]), user=user)
                self.assertEqual(response.status_code, 403)
        self.assertEqual(Expense.objects.count(), 0)

        rh = make_user("rh.admin", Role.ADMIN)
        response = self._importer(self._classeur([self._ligne()]), user=rh)

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(Expense.objects.count(), 1)

    def test_tout_arrive_en_brouillon(self):
        response = self._importer(self._classeur([self._ligne(STATUT="Justifié")]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Expense.objects.get().status, Status.DRAFT)
        self.assertEqual(Dossier.objects.get(number="N-IMPORT-01").status, Status.DRAFT)

    def test_le_montant_justifie_du_classeur_est_ignore(self):
        """Le pays déclare, le siège constate : un montant justifié ne
        s'importe pas, même s'il figure dans le fichier."""
        response = self._importer(
            self._classeur([self._ligne(**{"MONTANT JUSTIFIER": 1200})])
        )

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(str(Expense.objects.get().justified_amount), "0.00")

    # -- Idempotence --------------------------------------------------------

    def test_reimporter_le_meme_classeur_ne_cree_rien(self):
        classeur = self._classeur([self._ligne(), self._ligne(**{"LIBELLE DES TRANSACTIONS": "Hôtel"})])
        self._importer(classeur)
        classeur.seek(0)

        response = self._importer(classeur)

        self.assertEqual(len(response.data["erreurs"]), 2)
        self.assertIn("déjà présente", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 2)

    def test_une_ligne_identique_dans_le_meme_classeur_est_refusee(self):
        response = self._importer(self._classeur([self._ligne(), self._ligne()]))

        self.assertEqual(len(response.data["erreurs"]), 1)
        self.assertEqual(response.data["erreurs"][0]["ligne"], 3)
        self.assertIn("ligne 2", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_une_ligne_differente_s_ajoute_a_un_dossier_existant(self):
        self._importer(self._classeur([self._ligne()]))

        response = self._importer(self._classeur([self._ligne(DEPENSES=900)]))

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(response.data["dossiers_crees"], 0)
        self.assertEqual(Expense.objects.count(), 2)

    # -- Managers -----------------------------------------------------------

    def test_le_manager_est_resolu_dans_le_pays_de_la_ligne(self):
        """Deux homonymes dans deux pays : la ligne prend celui de son pays."""
        homonyme = Manager.objects.create(name="Kodjo Mensah")
        self.ivoire.managers.add(homonyme)

        response = self._importer(self._classeur([self._ligne()]))

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(Expense.objects.get().owner, self.manager)

    def test_un_homonyme_d_un_autre_pays_n_est_pas_reutilise(self):
        """Le manager ivoirien n'est pas rattaché au Togo en douce : un
        manager togolais est créé, le voisin reste tel quel."""
        voisin = Manager.objects.create(name="Awa Diop")
        self.ivoire.managers.add(voisin)

        response = self._importer(self._classeur([self._ligne(OWNER="Awa Diop")]))

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(response.data["managers_crees"], 1)
        owner = Expense.objects.get().owner
        self.assertNotEqual(owner, voisin)
        self.assertEqual(list(owner.countries.all()), [self.togo])
        self.assertEqual(list(voisin.countries.all()), [self.ivoire])

    def test_un_manager_inconnu_est_cree_et_rattache_au_pays(self):
        """Le classeur historique est la première source du référentiel :
        exiger sa saisie préalable rendrait l'import inutile. La création
        passe par le modèle, donc par l'historique."""
        response = self._importer(
            self._classeur([
                self._ligne(OWNER="Afi Lawson"),
                self._ligne(OWNER="afi lawson", **{"LIBELLE DES TRANSACTIONS": "Hôtel"}),
            ])
        )

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(response.data["managers_crees"], 1)
        cree = Manager.objects.get(name="Afi Lawson")
        self.assertEqual(list(cree.countries.all()), [self.togo])
        self.assertEqual(Expense.objects.filter(owner=cree).count(), 2)
        self.assertTrue(
            ChangeLog.objects.filter(
                model_name=ChangeLog.Models.MANAGER, action=ChangeLog.Actions.CREATED,
                performed_by=self.doo.username,
            ).exists()
        )

    def test_une_equipe_inconnue_est_creee_dans_le_pays(self):
        response = self._importer(self._classeur([self._ligne(TEAM="Équipe Kara")]))

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(response.data["equipes_creees"], 1)
        equipe = Team.objects.get(name="Équipe Kara")
        self.assertEqual(equipe.country, self.togo)
        self.assertEqual(Expense.objects.get().team, equipe)

    def test_la_previsualisation_ne_cree_ni_equipe_ni_manager(self):
        response = self._importer(
            self._classeur([self._ligne(TEAM="Équipe Kara", OWNER="Afi Lawson")]),
            dry_run="true",
        )

        self.assertEqual(response.data["equipes_creees"], 1)
        self.assertEqual(response.data["managers_crees"], 1)
        self.assertFalse(Team.objects.filter(name="Équipe Kara").exists())
        self.assertFalse(Manager.objects.filter(name="Afi Lawson").exists())

    # -- Devise d'origine (§5.3) -------------------------------------------

    def test_un_decaissement_en_devise_est_converti_au_taux_fige(self):
        ExchangeRate.objects.create(
            currency="EUR", rate_to_xof=Decimal("655.957"), valid_from=date(2020, 1, 1)
        )

        response = self._importer(
            self._classeur([
                self._ligne(**{"DEVISE D'ORIGINE": "eur", "MONTANT D'ORIGINE": 100})
            ])
        )

        self.assertFalse(response.data["erreurs"])
        expense = Expense.objects.get()
        self.assertEqual(expense.original_currency, "EUR")
        self.assertEqual(str(expense.original_amount), "100.00")
        self.assertEqual(str(expense.amount), "65595.70")
        self.assertEqual(str(expense.original_rate), "655.957000")

    def test_devise_sans_montant_d_origine_refusee(self):
        response = self._importer(
            self._classeur([self._ligne(**{"DEVISE D'ORIGINE": "EUR"})])
        )

        self.assertIn("aucun des deux", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_devise_sans_taux_refusee(self):
        response = self._importer(
            self._classeur([
                self._ligne(**{"DEVISE D'ORIGINE": "GBP", "MONTANT D'ORIGINE": 10})
            ])
        )

        self.assertIn("Aucun taux", response.data["erreurs"][0]["motif"])

    # -- Garde-fous ---------------------------------------------------------

    @override_settings(MAX_PROOF_SIZE=16)
    def test_un_classeur_trop_volumineux_est_refuse(self):
        response = self._importer(self._classeur([self._ligne()]))

        self.assertEqual(response.status_code, 200)
        self.assertIn("volumineux", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_le_nombre_de_lignes_est_plafonne(self):
        lignes = [self._ligne(**{"N°ORDRE": f"N-{i}"}) for i in range(3)]

        with mock.patch.object(imports, "LIGNES_MAX", 2):
            response = self._importer(self._classeur(lignes))

        self.assertIn("2 lignes", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_les_longueurs_de_champs_sont_verifiees_ligne_par_ligne(self):
        response = self._importer(
            self._classeur([
                self._ligne(**{"N°ORDRE": "N" * 51}),
                self._ligne(**{"N°ORDRE": "N-2", "LIBELLE DES TRANSACTIONS": "L" * 251}),
                self._ligne(**{"N°ORDRE": "N-3", "DEVISE D'ORIGINE": "EURO", "MONTANT D'ORIGINE": 1}),
            ])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual([e["ligne"] for e in response.data["erreurs"]], [2, 3, 4])
        for erreur in response.data["erreurs"]:
            self.assertIn("trop long", erreur["motif"])

    def test_les_montants_aberrants_sont_des_erreurs_de_ligne(self):
        """« NaN » et « Infinity » sont des Decimal valides, pas des montants."""
        response = self._importer(
            self._classeur([
                self._ligne(**{"N°ORDRE": "N-1", "DEPENSES": "NaN"}),
                self._ligne(**{"N°ORDRE": "N-2", "DEPENSES": "Infinity"}),
                self._ligne(**{"N°ORDRE": "N-3", "DEPENSES": "1" + "0" * 14}),
                self._ligne(**{"N°ORDRE": "N-4", "DEPENSES": "-5"}),
            ])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual([e["ligne"] for e in response.data["erreurs"]], [2, 3, 4, 5])
        self.assertEqual(Expense.objects.count(), 0)

    def test_le_meme_numero_dans_un_autre_pays_ne_gene_pas(self):
        """Le N°ORDRE est unique par pays : le « N-VOISIN » ivoirien
        n'empêche pas le Togo d'ouvrir le sien, et n'y reçoit aucune ligne."""
        voisin = Dossier.objects.create(
            number="N-VOISIN", label="Abidjan", country=self.ivoire,
            date=date(self.year, 1, 10),
        )

        response = self._importer(self._classeur([self._ligne(**{"N°ORDRE": "N-VOISIN"})]))

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(response.data["dossiers_crees"], 1)
        togolais = Dossier.objects.get(number="N-VOISIN", country=self.togo)
        self.assertEqual(togolais.expenses.count(), 1)
        self.assertEqual(voisin.expenses.count(), 0)

    def test_le_lecteur_xml_protege_est_utilise_quand_il_est_installe(self):
        """openpyxl bascule sur defusedxml dès que le paquet est présent : le
        drapeau du module doit dire la même chose que lui."""
        from openpyxl.xml import DEFUSEDXML

        self.assertEqual(imports.DEFUSEDXML_DISPONIBLE, DEFUSEDXML)

    # -- Écriture -----------------------------------------------------------

    def test_les_lignes_sont_inserees_par_lots(self):
        lignes = [
            self._ligne(**{"LIBELLE DES TRANSACTIONS": f"Ligne {i}"}) for i in range(30)
        ]

        with CaptureQueriesContext(connection) as captured:
            response = self._importer(self._classeur(lignes))

        self.assertFalse(response.data["erreurs"])
        insertions = [
            q["sql"] for q in captured.captured_queries
            if 'INSERT INTO "expenses_expense"' in q["sql"]
        ]
        self.assertEqual(len(insertions), 1)
        self.assertEqual(Expense.objects.count(), 30)

    def test_l_import_laisse_une_trace_avec_l_adresse_du_client(self):
        self.login(self.doo)
        self.client.post(
            "/api/imports/expenses.xlsx",
            {"file": ("depenses.xlsx", self._classeur([self._ligne()]), XLSX)},
            format="multipart",
            REMOTE_ADDR="10.20.30.40",
            # Sans mandataire déclaré, un en-tête forgé n'est pas cru.
            HTTP_X_FORWARDED_FOR="1.2.3.4",
        )

        entry = AuditLog.objects.get(object_type="ExpenseImport")
        self.assertEqual(entry.action, AuditLog.Action.IMPORTED)
        self.assertEqual(entry.ip_address, "10.20.30.40")

    # -- Comportements conservés -------------------------------------------

    def test_un_pays_hors_perimetre_est_refuse(self):
        """Les administrateurs voient tout : par l'API, ce cas ne se
        présente plus. La fonction garde sa garde-fou — un appelant futur au
        périmètre restreint ne doit pas verser chez le voisin."""
        resultat = imports.importer_depenses(
            self._classeur([self._ligne(PAYS="Côte d'Ivoire")]), self.owner, dry_run=True
        )

        self.assertEqual(resultat["erreurs"][0]["ligne"], 2)
        self.assertIn("hors périmètre", resultat["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_un_pays_inconnu_est_signale_sans_rien_creer(self):
        response = self._importer(self._classeur([self._ligne(PAYS="Bénin inconnu")]))

        self.assertEqual(response.data["erreurs"][0]["ligne"], 2)
        self.assertIn("inconnu", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_la_previsualisation_n_ecrit_rien(self):
        avant = Expense.objects.count()
        response = self._importer(self._classeur([self._ligne()]), dry_run="true")

        self.assertEqual(response.data["dossiers_crees"], 1)
        self.assertEqual(response.data["lignes_creees"], 1)
        self.assertEqual(Expense.objects.count(), avant)

    def test_un_echec_partiel_n_ecrit_rien(self):
        response = self._importer(
            self._classeur([
                self._ligne(**{"N°ORDRE": "N-1"}),
                self._ligne(**{"N°ORDRE": "N-2", "DEPENSES": "12 000 F"}),
                self._ligne(**{"N°ORDRE": "N-3"}),
                self._ligne(**{"N°ORDRE": "N-4"}),
            ])
        )

        self.assertEqual(len(response.data["erreurs"]), 1)
        self.assertEqual(Expense.objects.count(), 0)
        self.assertEqual(Dossier.objects.count(), 1)

    def test_un_dossier_deja_declare_est_refuse(self):
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        response = self._importer(
            self._classeur([self._ligne(**{"N°ORDRE": self.dossier.number})])
        )

        self.assertIn("déclaré", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_les_colonnes_sont_retrouvees_par_leur_en_tete(self):
        entetes = [title for title, _ in EXPENSE_COLUMNS][::-1]
        response = self._importer(self._classeur([self._ligne()], entetes))

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(Expense.objects.get().title, "Déplacement")


class ClasseurHistoriqueTests(ExpenseTestCase):
    """Le classeur réel du client, reproduit anonymement.

    Une feuille « BASE DE DONNEES ACTIONS », un titre fusionné en ligne 2,
    une note en ligne 4, l'en-tête en ligne 7 et neuf colonnes : N°ORDRE,
    DATE, TEAM, OWNER, LIBELLE DES TRANSACTIONS, DEPENSES, MONTANT JUSTIFIER,
    ECART, PIECES JUSTIFICATIVES. Les N°ORDRE sont des entiers numérotés par
    pays, un même numéro regroupant plusieurs lignes ; les dates n'ont pas
    d'heure ; les montants sont entiers ; MONTANT JUSTIFIER est parfois vide.
    """

    COLONNES = [
        "N°ORDRE", "DATE", "TEAM", "OWNER", "LIBELLE DES TRANSACTIONS",
        "DEPENSES", "MONTANT JUSTIFIER", "ECART", "PIECES JUSTIFICATIVES",
    ]

    def setUp(self):
        super().setUp()
        self.lignes = [
            [1, datetime(self.year, 1, 6), "Équipe A", "Owner Un", "Carburant", 15000, 15000, 0, "Reçu"],
            [1, datetime(self.year, 1, 6), "Équipe A", "Owner Un", "Péage", 2000, None, 2000, ""],
            [1, datetime(self.year, 1, 7), "Équipe A", "Owner Un", "Hôtel", 45000, 30000, 15000, "Reçu(justif incomplet)"],
            [2, datetime(self.year, 1, 12), "Équipe B", "Owner Deux", "Impression flyers", 80000, 80000, 0, "Reçu"],
        ]

    def _classeur(self, lignes=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BASE DE DONNEES ACTIONS"
        sheet.merge_cells("A2:I2")
        sheet["A2"] = "BASE DE DONNÉES DES ACTIONS — 1er trimestre"
        sheet["A4"] = "Note : les montants sont en francs, les pièces au bureau."
        for colonne, entete in enumerate(self.COLONNES, start=1):
            sheet.cell(row=7, column=colonne, value=entete)
        for indice, ligne in enumerate(lignes if lignes is not None else self.lignes, start=8):
            for colonne, valeur in enumerate(ligne, start=1):
                sheet.cell(row=indice, column=colonne, value=valeur)
        contenu = BytesIO()
        workbook.save(contenu)
        contenu.seek(0)
        return contenu

    def _importer(self, contenu, user=None, **query):
        self.login(user or self.doo)
        url = "/api/imports/expenses.xlsx"
        if query:
            url += "?" + urlencode(query)
        return self.client.post(
            url, {"file": ("historique.xlsx", contenu, XLSX)}, format="multipart"
        )

    def test_le_classeur_historique_s_importe_dans_le_pays_de_la_requete(self):
        response = self._importer(self._classeur(), country=self.togo.pk)

        self.assertEqual(response.status_code, 200, response.data)
        self.assertFalse(response.data["erreurs"])
        self.assertEqual(response.data["dossiers_crees"], 2)
        self.assertEqual(response.data["lignes_creees"], 4)
        premier = Dossier.objects.get(country=self.togo, number="1")
        self.assertEqual(premier.expenses.count(), 3)
        self.assertEqual(premier.status, Status.DRAFT)
        self.assertEqual(premier.date, date(self.year, 1, 6))
        second = Dossier.objects.get(country=self.togo, number="2")
        self.assertEqual(second.expenses.count(), 1)
        self.assertTrue(
            all(e.country == self.togo for e in Expense.objects.all())
        )

    def test_les_dates_sans_heure_et_les_montants_entiers_sont_lus(self):
        self._importer(self._classeur(), country=self.togo.pk)

        hotel = Expense.objects.get(title="Hôtel")
        self.assertEqual(hotel.date.date(), date(self.year, 1, 7))
        self.assertEqual(str(hotel.amount), "45000.00")

    def test_le_montant_justifie_et_l_ecart_sont_ignores(self):
        """Le siège constate : le classeur peut dire « justifié », il ne
        prouve rien. Une cellule vide n'est pas non plus une erreur."""
        self._importer(self._classeur(), country=self.togo.pk)

        self.assertEqual(
            set(Expense.objects.values_list("justified_amount", flat=True)),
            {Decimal("0.00")},
        )

    def test_la_mention_de_piece_est_conservee_en_remarque(self):
        self._importer(self._classeur(), country=self.togo.pk)

        self.assertEqual(Expense.objects.get(title="Hôtel").note, "Pièce : Reçu(justif incomplet)")
        self.assertEqual(Expense.objects.get(title="Carburant").note, "Pièce : Reçu")
        self.assertEqual(Expense.objects.get(title="Péage").note, "")

    def test_equipes_et_managers_inconnus_sont_crees_dans_le_pays(self):
        response = self._importer(self._classeur(), country=self.togo.pk)

        self.assertEqual(response.data["equipes_creees"], 2)
        self.assertEqual(response.data["managers_crees"], 2)
        equipe = Team.objects.get(name="Équipe A")
        self.assertEqual(equipe.country, self.togo)
        owner = Manager.objects.get(name="Owner Un")
        self.assertEqual(list(owner.countries.all()), [self.togo])
        self.assertEqual(Expense.objects.filter(team=equipe, owner=owner).count(), 3)

    def test_sans_colonne_pays_le_parametre_country_est_obligatoire(self):
        response = self._importer(self._classeur())

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["erreurs"][0]["ligne"], 1)
        self.assertIn("PAYS", response.data["erreurs"][0]["motif"])
        self.assertIn("country", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)
        self.assertEqual(Team.objects.filter(name="Équipe A").count(), 0)

    def test_un_pays_inconnu_est_refuse(self):
        inconnu = self._importer(self._classeur(), country=self.ivoire.pk + 1000)

        self.assertEqual(inconnu.status_code, 400)
        self.assertIn("country", inconnu.data)
        self.assertEqual(Expense.objects.count(), 0)

    def test_le_responsable_pays_ne_verse_pas_de_classeur_meme_chez_lui(self):
        """Le pays déclare dans l'application ; le classeur historique est
        repris par les administrateurs."""
        response = self._importer(self._classeur(), user=self.owner, country=self.togo.pk)

        self.assertEqual(response.status_code, 403)
        self.assertEqual(Expense.objects.count(), 0)

    def test_le_siege_importe_dans_n_importe_quel_pays(self):
        siege = make_user("ceo.innov", Role.SUPER_ADMIN)

        response = self._importer(self._classeur(), user=siege, country=self.ivoire.pk)

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(Dossier.objects.filter(country=self.ivoire).count(), 2)

    def test_le_pays_peut_venir_du_formulaire(self):
        self.login(self.doo)

        response = self.client.post(
            "/api/imports/expenses.xlsx",
            {"file": ("historique.xlsx", self._classeur(), XLSX), "country": self.togo.pk},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data["lignes_creees"], 4)

    def test_la_previsualisation_fonctionne_sur_ce_format(self):
        response = self._importer(self._classeur(), country=self.togo.pk, dry_run="true")

        self.assertTrue(response.data["dry_run"])
        self.assertEqual(response.data["dossiers_crees"], 2)
        self.assertEqual(response.data["lignes_creees"], 4)
        self.assertEqual(response.data["equipes_creees"], 2)
        self.assertEqual(response.data["managers_crees"], 2)
        self.assertEqual(Expense.objects.count(), 0)
        self.assertEqual(Dossier.objects.filter(country=self.togo).count(), 1)
        self.assertFalse(Team.objects.filter(name="Équipe A").exists())
        self.assertFalse(Manager.objects.filter(name="Owner Un").exists())

    def test_les_erreurs_sont_signalees_au_numero_de_ligne_du_classeur(self):
        """La ligne 9 du classeur est la ligne 9, pas la deuxième après
        l'en-tête : le déclarant doit la retrouver dans Excel."""
        lignes = list(self.lignes)
        lignes[1] = [1, datetime(self.year, 1, 6), "Équipe A", "Owner Un", "Péage", "deux mille", None, None, ""]

        response = self._importer(self._classeur(lignes), country=self.togo.pk)

        self.assertEqual(len(response.data["erreurs"]), 1)
        self.assertEqual(response.data["erreurs"][0]["ligne"], 9)
        self.assertIn("DEPENSES", response.data["erreurs"][0]["motif"])
        self.assertEqual(Expense.objects.count(), 0)

    def test_un_numero_d_ordre_numerique_rejoint_le_dossier_saisi_a_la_main(self):
        """Un N°ORDRE lu en flottant donnerait « 1.0 » : il doit rejoindre
        le dossier « 1 » ouvert à la main."""
        self.dossier.number = "1"
        self.dossier.save()
        lignes = [[1.0, datetime(self.year, 1, 6), "Équipe A", "Owner Un", "Carburant", 15000, 15000, 0, "Reçu"]]

        response = self._importer(self._classeur(lignes), country=self.togo.pk)

        self.assertFalse(response.data["erreurs"])
        self.assertEqual(response.data["dossiers_crees"], 0)
        self.assertEqual(self.dossier.expenses.count(), 1)

    def test_reimporter_le_classeur_historique_ne_cree_rien(self):
        self._importer(self._classeur(), country=self.togo.pk)

        response = self._importer(self._classeur(), country=self.togo.pk)

        self.assertEqual(len(response.data["erreurs"]), 4)
        self.assertEqual(Expense.objects.count(), 4)
        self.assertEqual(Team.objects.filter(name="Équipe A").count(), 1)

    def test_les_lignes_importees_se_soumettent_ensuite(self):
        """L'import fournit équipe et manager : le dossier peut partir."""
        self._importer(self._classeur(), country=self.togo.pk)
        premier = Dossier.objects.get(country=self.togo, number="1")

        response = self.submit_dossier(premier)

        self.assertEqual(response.status_code, 200, response.data)

    def test_l_import_est_journalise_avec_son_pays(self):
        self._importer(self._classeur(), country=self.togo.pk)

        entry = AuditLog.objects.get(object_type="ExpenseImport")
        self.assertEqual(entry.country, self.togo)
        self.assertEqual(entry.detail["lignes_creees"], 4)

    def test_sans_ligne_d_entete_le_classeur_est_refuse(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Un classeur qui n'a rien à voir"
        sheet["A2"] = 42
        contenu = BytesIO()
        workbook.save(contenu)
        contenu.seek(0)

        response = self._importer(contenu, country=self.togo.pk)

        self.assertEqual(response.data["erreurs"][0]["ligne"], 1)
        self.assertIn("introuvable", response.data["erreurs"][0]["motif"])


class ImportsConcurrentsTests(ImportTests):
    """Deux imports du même classeur peuvent se croiser : le second ne
    répond jamais 500, il refuse la ligne en la nommant, sans rien écrire."""

    def _autre_import_cree_le_dossier(self):
        """Simule l'autre import, validé en même temps et écrit juste avant."""
        original = Dossier.objects.get_or_create

        def concurrent(**kwargs):
            Dossier.objects.create(
                country=kwargs["country"], number=kwargs["number"],
                label="Créé par l'autre import", date=date(self.year, 1, 1),
                created_by="autre.import",
            )
            return original(**kwargs)

        return mock.patch.object(Dossier.objects, "get_or_create", side_effect=concurrent)

    def test_un_dossier_cree_entre_temps_est_une_erreur_de_ligne(self):
        with self._autre_import_cree_le_dossier():
            response = self._importer(self._classeur([self._ligne()]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["lignes_creees"], 0)
        self.assertEqual(response.data["erreurs"][0]["ligne"], 2)
        self.assertIn("N-IMPORT-01", response.data["erreurs"][0]["motif"])
        self.assertIn("autre import", response.data["erreurs"][0]["motif"])
        # Rien de cet import n'est écrit. (L'autre import est simulé dans la
        # même transaction : son dossier disparaît avec le point de reprise,
        # alors qu'en réalité il est déjà validé — c'est tout l'objet du test.)
        self.assertEqual(Expense.objects.count(), 0)
        self.assertTrue(
            AuditLog.objects.filter(action=AuditLog.Action.IMPORTED).exists()
        )

    def test_une_violation_d_unicite_est_une_erreur_de_ligne(self):
        with mock.patch.object(
            Dossier.objects, "get_or_create", side_effect=IntegrityError("doublon")
        ):
            response = self._importer(self._classeur([self._ligne()]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["erreurs"][0]["ligne"], 2)
        self.assertEqual(Dossier.objects.count(), 1)
        self.assertEqual(Expense.objects.count(), 0)


class ClasseurGonfleTests(ExpenseTestCase):
    """Un xlsx est une archive : quelques Ko compressés peuvent en cacher
    des centaines de Mo, qu'openpyxl chargerait en mémoire."""

    def test_un_classeur_trop_gros_une_fois_decompresse_est_refuse(self):
        import io
        import zipfile

        from django.conf import settings
        from django.core.files.uploadedfile import SimpleUploadedFile
        from rest_framework import status

        tampon = io.BytesIO()
        with zipfile.ZipFile(tampon, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", "<Types/>")
            archive.writestr("xl/sharedStrings.xml", b"\x00" * (5 * settings.MAX_PROOF_SIZE + 1))
        self.assertLess(tampon.tell(), 1024 * 1024)
        self.login(self.doo)

        response = self.client.post(
            "/api/imports/expenses.xlsx",
            {"file": SimpleUploadedFile("bombe.xlsx", tampon.getvalue(), content_type=XLSX)},
            format="multipart",
        )

        # Comme tout classeur illisible : rien n'est écrit, le motif est rendu.
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["lignes_creees"], 0)
        self.assertIn("décompressé", response.data["erreurs"][0]["motif"])


class ReimportDUnExportTests(ExpenseTestCase):
    """Un classeur exporté se réimporte sans rien créer, même quand les
    lignes portent une heure : le classeur ne connaît que le jour."""

    def test_une_ligne_saisie_a_neuf_heures_n_est_pas_recreee(self):
        from datetime import datetime, timezone as tz

        self.make_expense(amount="1200.00", date=datetime(self.year, 3, 15, 9, 0, tzinfo=tz.utc), title="Carburant")
        self.login(self.doo)
        classeur = self.client.get(f"/api/exports/expenses.xlsx?year={self.year}&country={self.togo.pk}").content

        response = self.client.post(
            "/api/imports/expenses.xlsx",
            {"file": SimpleUploadedFile("export.xlsx", classeur, content_type=XLSX), "country": self.togo.pk},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["lignes_creees"], 0, response.data)
        self.assertEqual(self.dossier.expenses.count(), 1)
