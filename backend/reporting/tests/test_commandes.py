"""Commandes planifiées et d'amorçage.

Les tests des commandes vivent ici, avec ceux du pilotage qu'elles servent :
ordonnanceur, alertes, rapport périodique, amorçage des comptes, bucket.
"""

import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import mock

from django.contrib.auth.models import User
from django.core import mail
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import Role
from accounts.tests.test_scoping import make_user
from budget.models import Budget
from core.management.commands.run_scheduler import JOBS, declencheur, run_job
from core.models import ChangeLog, Country
from expenses.models import Dossier, Proof
from expenses.tests.base import in_memory_storage
from expenses.workflow import Status
from notifications.models import Notification
from reporting.tests.test_dashboard import DashboardTestCase


class OrdonnanceurTests(TestCase):
    def test_la_connexion_est_rafraichie_avant_et_apres_chaque_tache(self):
        """L'ordonnanceur vit des jours : une connexion fermée par le serveur
        faisait échouer la tâche suivante sur un descripteur mort."""
        with mock.patch(
            "core.management.commands.run_scheduler.close_old_connections"
        ) as fermer, mock.patch(
            "core.management.commands.run_scheduler.call_command"
        ) as appel:
            run_job(JOBS[0])

        self.assertEqual(fermer.call_count, 2)
        appel.assert_called_once()

    def test_la_connexion_est_rafraichie_meme_si_la_tache_echoue(self):
        with mock.patch(
            "core.management.commands.run_scheduler.close_old_connections"
        ) as fermer, mock.patch(
            "core.management.commands.run_scheduler.call_command",
            side_effect=OSError("SMTP injoignable"),
        ), self.assertLogs("scheduler", level="ERROR"):
            run_job(JOBS[0])

        self.assertEqual(fermer.call_count, 2)

    def test_une_cadence_invalide_est_nommee(self):
        """Une expression fautive dans l'environnement doit dire quelle
        variable corriger, pas une trace apscheduler."""
        with self.assertRaises(CommandError) as capture:
            declencheur(JOBS[0], "chaque lundi")

        self.assertIn("SCHEDULE_ALERTS", str(capture.exception))
        self.assertIn("chaque lundi", str(capture.exception))

    def test_une_cadence_valide_donne_un_declencheur(self):
        self.assertIsNotNone(declencheur(JOBS[0], "*/5 * * * *"))


class NotifyAlertsTests(DashboardTestCase):
    dossier_status = Status.SUBMITTED

    def test_une_enveloppe_desactivee_ne_notifie_pas(self):
        Budget.objects.filter(pk=self.budget.pk).update(
            amount=Decimal("100000.00"), is_active=False
        )

        call_command("notify_alerts", year=self.year, verbosity=0)

        self.assertFalse(
            Notification.objects.filter(kind=Notification.Kind.BUDGET_OVERRUN).exists()
        )

    def test_en_janvier_les_dossiers_de_decembre_sont_encore_rappeles(self):
        """Borner à l'année civile ferait taire le rappel au moment où il
        compte : un dossier de décembre sans preuve presse en janvier."""
        ancien = Dossier.objects.create(
            number="N-DEC", label="Fin d'exercice", country=self.togo,
            date=date(self.year - 1, 12, 20), status=Status.SUBMITTED,
        )
        janvier = timezone.make_aware(datetime(self.year, 1, 15, 9, 0))

        with mock.patch("django.utils.timezone.now", return_value=janvier):
            call_command("notify_alerts", verbosity=0)

        self.assertTrue(
            Notification.objects.filter(
                recipient=self.controller, link=f"/dossiers/{ancien.pk}"
            ).exists()
        )

    def test_une_annee_explicite_reste_bornee(self):
        ancien = Dossier.objects.create(
            number="N-DEC", label="Fin d'exercice", country=self.togo,
            date=date(self.year - 1, 12, 20), status=Status.SUBMITTED,
        )

        call_command("notify_alerts", year=self.year, verbosity=0)

        self.assertFalse(
            Notification.objects.filter(link=f"/dossiers/{ancien.pk}").exists()
        )


@in_memory_storage
class RapportPeriodiqueTests(DashboardTestCase):
    def setUp(self):
        super().setUp()
        self.doo.email = "doo@example.org"
        self.doo.save()
        self.controle_togo = make_user("controle.togo", Role.DF, [self.togo])
        self.controle_togo.email = "togo@example.org"
        self.controle_togo.save()
        self.abidjan = Dossier.objects.create(
            number="CI-0001", label="Salon Abidjan", country=self.ivoire,
            date=date(self.year, 4, 2), status=Status.SUBMITTED,
        )

    def _classeur(self, message):
        _, contenu, _ = message.attachments[0]
        return load_workbook(BytesIO(contenu))["Rapprochement dossiers"]

    def _numeros(self, message):
        # La ligne TOTAL n'a pas de numéro.
        return {
            row[0] for row in self._classeur(message).iter_rows(min_row=2, values_only=True)
        } - {None}

    def test_une_direction_financiere_restreinte_ne_recoit_que_son_perimetre(self):
        """Régression : la synthèse entière partait à tous, dossiers du
        voisin compris — le cloisonnement contourné par un e-mail."""
        call_command("send_periodic_report", year=self.year, verbosity=0)

        par_destinataire = {tuple(m.bcc): m for m in mail.outbox}
        togo = par_destinataire[("togo@example.org",)]
        siege = par_destinataire[("doo@example.org",)]
        self.assertIn("Togo", togo.body)
        self.assertNotIn("CI-0001", togo.body)
        self.assertNotIn("Côte d'Ivoire", togo.body)
        self.assertEqual(self._numeros(siege), {"N-0001", "CI-0001"})

    def test_le_dm_recoit_le_rapport_de_son_perimetre(self):
        """Le DM met en contrôle : il lit la même synthèse que le DF, sur son
        périmètre, sans classeur."""
        dm_togo = make_user("dm.togo", Role.DM, [self.togo], email="dm.togo@example.org")

        call_command("send_periodic_report", year=self.year, verbosity=0)

        # Même périmètre, même langue, pas de classeur : le DM du Togo lit
        # le message du DF du Togo, en copie cachée avec lui.
        message = next(m for m in mail.outbox if "dm.togo@example.org" in m.bcc)
        self.assertIn("Togo", message.body)
        self.assertNotIn("CI-0001", message.body)
        self.assertEqual(message.attachments, [])
        self.assertFalse(
            any(m.bcc == [self.owner.email] for m in mail.outbox),
            "le manager ne reçoit pas le rapport du siège",
        )
        self.assertEqual(dm_togo.profile.role, Role.DM)

    def test_la_piece_jointe_n_est_envoyee_qu_aux_administrateurs(self):
        """Seuls les administrateurs manipulent des fichiers : la direction
        financière lit la synthèse et retrouve le détail dans l'application."""
        call_command("send_periodic_report", year=self.year, verbosity=0)

        par_destinataire = {tuple(m.bcc): m for m in mail.outbox}
        togo = par_destinataire[("togo@example.org",)]
        siege = par_destinataire[("doo@example.org",)]
        self.assertEqual(togo.attachments, [])
        self.assertIn("dans l'application", togo.body)
        self.assertEqual(len(siege.attachments), 1)
        self.assertIn("en pièce jointe", siege.body)

    def test_le_rapport_est_dans_la_langue_du_destinataire(self):
        with mock.patch(
            "reporting.management.commands.send_periodic_report.langue_de",
            side_effect=lambda user: "en" if user == self.controle_togo else "fr",
        ):
            call_command("send_periodic_report", year=self.year, verbosity=0)

        par_destinataire = {tuple(m.bcc): m for m in mail.outbox}
        togo = par_destinataire[("togo@example.org",)]
        siege = par_destinataire[("doo@example.org",)]
        self.assertEqual(togo.subject, f"[Budget control] weekly report — {self.year}")
        self.assertIn("Dossiers without any supporting document", togo.body)
        self.assertEqual(siege.subject, f"[Contrôle budgétaire] Rapport hebdomadaire — {self.year}")
        self.assertIn("Dossiers sans aucun justificatif", siege.body)

    def test_les_adresses_ne_sont_pas_exposees(self):
        for message in mail.outbox:
            self.assertEqual(message.to, [])
        call_command("send_periodic_report", year=self.year, verbosity=0)
        for message in mail.outbox:
            self.assertEqual(message.to, [])
            self.assertEqual(len(message.bcc), 1)

    def test_les_dossiers_sans_justificatif_suivent_la_regle_des_alertes(self):
        """Un brouillon ne compte pas ; une pièce rejetée ne couvre rien."""
        brouillon = Dossier.objects.create(
            number="N-BROUILLON", label="Pas encore déclaré", country=self.togo,
            date=date(self.year, 5, 1), status=Status.DRAFT,
        )
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        Proof.objects.create(
            dossier=self.dossier, file="justificatifs/rejete.pdf",
            original_name="rejete.pdf", kind=Proof.Kind.INVOICE,
            status=Proof.ProofStatus.REJECTED, sha256="b" * 64,
        )
        self.assertEqual(brouillon.proofs.count(), 0)

        call_command("send_periodic_report", year=self.year, verbosity=0)

        siege = next(m for m in mail.outbox if m.bcc == ["doo@example.org"])
        # N-0001 (pièce rejetée) et CI-0001 (aucune pièce) ; pas le brouillon.
        self.assertIn("Dossiers sans aucun justificatif : 2", siege.body)

    def test_les_devises_ne_s_additionnent_qu_en_fcfa(self):
        ghana = Country.objects.create(
            name="Guinée", code="GN", country_ref="GN-03", currency="GNF",
            timezone="Africa/Conakry",
        )
        Budget.objects.create(country=ghana, year=self.year, amount=Decimal("5000.00"))

        call_command("send_periodic_report", year=self.year, verbosity=0)

        siege = next(m for m in mail.outbox if m.bcc == ["doo@example.org"])
        self.assertIn("Guinée (GNF) : attribué 5000.00", siege.body)
        self.assertIn("Enveloppes attribuées : 1500000.00", siege.body)
        self.assertIn("faute de taux : GNF", siege.body)

    def test_sans_destinataire_rien_ne_part(self):
        User.objects.update(email="")
        sortie = StringIO()

        call_command("send_periodic_report", year=self.year, stdout=sortie)

        self.assertEqual(mail.outbox, [])
        self.assertIn("Aucun destinataire", sortie.getvalue())


class SeedUsersTests(TestCase):
    def setUp(self):
        self.dossier_temp = TemporaryDirectory()
        self.fichier = Path(self.dossier_temp.name) / "seed.json"

    def tearDown(self):
        self.dossier_temp.cleanup()

    def _ecrire(self, users, countries=None):
        countries = countries if countries is not None else [{
            "country_ref": "TG-02", "name": "Togo", "code": "TG",
            "currency": "XOF", "timezone": "Africa/Lome",
        }]
        self.fichier.write_text(
            json.dumps({"countries": countries, "users": users}), encoding="utf-8"
        )

    def _lancer(self):
        sortie = StringIO()
        call_command("seed_users", file=str(self.fichier), stdout=sortie)
        return sortie.getvalue()

    def _compte(self, **extra):
        username = extra.pop("username", "pays.innov")
        # Chaque compte porte une adresse professionnelle : c'est elle qui
        # reçoit les notifications, et la commande la refuse absente.
        return {
            "username": username, "password": "Secret-Provisoire-2026",
            "email": f"{username}@innovpharma.net",
            "role": "manager", "countries": ["TG-02"], **extra,
        }

    def test_creation_puis_relance_sans_effet(self):
        self._ecrire([self._compte()])
        self._lancer()
        user = User.objects.get(username="pays.innov")
        self.assertTrue(user.check_password("Secret-Provisoire-2026"))
        self.assertTrue(user.profile.must_change_password)

        # Le titulaire remplace son mot de passe provisoire…
        user.set_password("Choisi-Par-Moi-2026")
        user.save()
        user.profile.must_change_password = False
        user.profile.save()
        # …et la commande est relancée pour ajouter un autre compte.
        self._ecrire([self._compte(), self._compte(username="autre.innov")])
        self._lancer()

        user.refresh_from_db()
        self.assertTrue(user.check_password("Choisi-Par-Moi-2026"))
        self.assertFalse(user.profile.must_change_password)
        self.assertTrue(User.objects.filter(username="autre.innov").exists())

    def test_reset_password_explicite_repose_le_mot_de_passe(self):
        self._ecrire([self._compte()])
        self._lancer()
        user = User.objects.get(username="pays.innov")
        user.set_password("Choisi-Par-Moi-2026")
        user.save()
        user.profile.must_change_password = False
        user.profile.save()

        self._ecrire([self._compte(password="Nouveau-2026", reset_password=True)])
        self._lancer()

        user.refresh_from_db()
        self.assertTrue(user.check_password("Nouveau-2026"))
        self.assertTrue(user.profile.must_change_password)

    def test_le_mot_de_passe_n_est_jamais_journalise(self):
        self._ecrire([self._compte(password="Ultra-Secret-2026")])

        sortie = self._lancer()

        self.assertNotIn("Ultra-Secret-2026", sortie)
        self.assertIn("pays.innov", sortie)

    def test_un_pays_hors_afrique_est_refuse_clairement(self):
        self._ecrire([], countries=[{
            "country_ref": "FR-99", "name": "France", "code": "FR",
            "currency": "EUR", "timezone": "Europe/Paris",
        }])

        with self.assertRaises(CommandError) as capture:
            self._lancer()

        self.assertIn("FR-99", str(capture.exception))
        self.assertIn("code", str(capture.exception))
        self.assertFalse(Country.objects.filter(code="FR").exists())

    def test_un_manager_sans_pays_est_refuse(self):
        self._ecrire([self._compte(countries=[])])

        with self.assertRaises(CommandError) as capture:
            self._lancer()

        self.assertIn("au moins un pays", str(capture.exception))

    def test_l_historique_est_signe_par_la_commande(self):
        self._ecrire([])

        self._lancer()

        entree = ChangeLog.objects.filter(model_name=ChangeLog.Models.COUNTRY).first()
        self.assertIsNotNone(entree)
        self.assertEqual(entree.performed_by, "seed_users")


@override_settings(AWS_S3_ENDPOINT_URL="http://minio:9000")
class EnsureBucketTests(TestCase):
    def _erreur(self, code, statut):
        from botocore.exceptions import ClientError

        return ClientError(
            {"Error": {"Code": code}, "ResponseMetadata": {"HTTPStatusCode": statut}},
            "HeadBucket",
        )

    def test_un_bucket_absent_est_cree(self):
        with mock.patch("boto3.client") as fabrique:
            client = fabrique.return_value
            client.head_bucket.side_effect = self._erreur("404", 404)
            call_command("ensure_bucket", stdout=StringIO())

        client.create_bucket.assert_called_once()

    def test_des_identifiants_faux_sont_nommes_sans_tenter_la_creation(self):
        with mock.patch("boto3.client") as fabrique:
            client = fabrique.return_value
            client.head_bucket.side_effect = self._erreur("403", 403)
            with self.assertRaises(CommandError) as capture:
                call_command("ensure_bucket", stdout=StringIO())

        self.assertIn("Accès refusé", str(capture.exception))
        client.create_bucket.assert_not_called()

    def test_un_bucket_present_est_laisse_tel_quel(self):
        sortie = StringIO()
        with mock.patch("boto3.client") as fabrique:
            call_command("ensure_bucket", stdout=sortie)

        fabrique.return_value.create_bucket.assert_not_called()
        self.assertIn("déjà présent", sortie.getvalue())
