"""Tableaux de bord, alertes, notifications et exports."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.core import mail
from django.core.management import call_command
from django.utils import translation
from openpyxl import load_workbook
from rest_framework import status

from accounts.models import Role
from accounts.tests.test_scoping import make_user
from budget.models import Budget, ExchangeRate
from core.models import Country, Project, Team
from expenses.models import AuditLog, Dossier, Expense, Proof
from expenses.tests.base import ExpenseTestCase, in_memory_storage
from expenses.workflow import Status
from notifications.models import Notification


class DashboardTestCase(ExpenseTestCase):
    """Une dépense justifiée et une soumise, sur l'enveloppe togolaise."""

    def setUp(self):
        super().setUp()
        self.justifiee = self.make_expense(
            amount="300000.00", justified_amount="250000.00",
            status=Status.JUSTIFIED, budget=self.budget,
        )
        self.soumise = self.make_expense(
            amount="200000.00", status=Status.SUBMITTED, budget=self.budget
        )


class DashboardTests(DashboardTestCase):
    def test_consolidation_et_indicateurs(self):
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        totals = response.data["totals"]
        self.assertEqual(totals["allocated"], "1500000.00")
        self.assertEqual(totals["consumed"], "300000.00")
        self.assertEqual(totals["engaged"], "200000.00")
        self.assertEqual(totals["justified"], "250000.00")
        # Le disponible retranche l'engagé comme le consommé.
        self.assertEqual(totals["remaining"], "1000000.00")

    def test_repartition_par_pays(self):
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        togo = next(
            row for row in response.data["countries"] if row["country_ref"] == "TG-02"
        )
        self.assertEqual(togo["remaining"], "500000.00")
        self.assertEqual(togo["justification_rate"], "0.8333")

    def test_charge_de_travail(self):
        self.login(self.controller)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        self.assertEqual(response.data["workload"]["expenses_to_review"], 1)

    def test_un_pays_ne_voit_que_ses_chiffres(self):
        self.login(self.rep_ivoire)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        refs = [row["country_ref"] for row in response.data["countries"]]
        self.assertEqual(refs, ["CT-01"])
        self.assertEqual(response.data["totals"]["consumed"], "0.00")

    def test_un_pays_hors_perimetre_repond_404(self):
        """Le pays du voisin n'existe pas : ni chiffres, ni confirmation."""
        self.login(self.rep_ivoire)

        response = self.client.get(
            "/api/dashboard/", {"year": self.year, "country": self.togo.pk}
        )

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_conversion_en_fcfa_signale_les_devises_inconnues(self):
        self.ivoire.currency = "GNF"
        self.ivoire.save()
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        self.assertEqual(
            response.data["consolidated_xof"]["unconverted_currencies"], ["GNF"]
        )
        # Le Togo reste converti : le total n'absorbe pas la devise inconnue.
        self.assertEqual(response.data["consolidated_xof"]["remaining"], "500000.00")

    def test_les_totaux_ne_melangent_pas_les_devises(self):
        """Additionner des francs et des cedis donnait un chiffre sans unité,
        présenté comme un total."""
        ghana = Country.objects.create(
            name="Guinée", code="GN", country_ref="GN-03", currency="GNF",
            timezone="Africa/Conakry",
        )
        Budget.objects.create(country=ghana, year=self.year, amount=Decimal("5000.00"))
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        totals = response.data["totals"]
        self.assertEqual(totals["currency"], "XOF")
        self.assertEqual(totals["allocated"], "1500000.00")
        self.assertEqual(totals["unconverted_currencies"], ["GNF"])
        ghana_row = next(r for r in response.data["countries"] if r["currency"] == "GNF")
        self.assertEqual(ghana_row["allocated"], "5000.00")

    def test_une_devise_convertie_entre_dans_les_totaux(self):
        ghana = Country.objects.create(
            name="Guinée", code="GN", country_ref="GN-03", currency="GNF",
            timezone="Africa/Conakry",
        )
        Budget.objects.create(country=ghana, year=self.year, amount=Decimal("5000.00"))
        ExchangeRate.objects.create(
            currency="GNF", rate_to_xof=Decimal("40"), valid_from=date(2020, 1, 1)
        )
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        self.assertEqual(response.data["totals"]["allocated"], "1700000.00")
        self.assertEqual(response.data["totals"]["unconverted_currencies"], [])

    def test_conversion_utilise_le_taux_enregistre(self):
        self.ivoire.currency = "EUR"
        self.ivoire.save()
        ExchangeRate.objects.create(
            currency="EUR", rate_to_xof=Decimal("655.957"), valid_from=date(2020, 1, 1)
        )
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        ivoire = next(
            r for r in response.data["countries"] if r["country_ref"] == "CT-01"
        )
        self.assertEqual(ivoire["remaining_xof"], "327978500.00")


class BreakdownTests(DashboardTestCase):
    def test_un_compte_restreint_a_un_pays_obtient_sa_repartition_sans_le_nommer(self):
        """Le pays est implicite quand le périmètre n'en contient qu'un."""
        self.login(self.owner)

        response = self.client.get("/api/dashboard/breakdown/", {"year": self.year})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        equipes = {row["label"] for row in response.data["by_team"]}
        self.assertIn("Équipe Lomé", equipes)

    def test_repartition_par_equipe_et_par_mois(self):
        self.login(self.doo)

        response = self.client.get(
            "/api/dashboard/breakdown/", {"year": self.year, "country": self.togo.pk}
        )

        equipes = {row["label"]: row for row in response.data["by_team"]}
        self.assertEqual(equipes["Équipe Lomé"]["amount"], "500000.00")
        self.assertEqual(equipes["Équipe Lomé"]["lines"], 2)
        self.assertEqual(len(response.data["by_month"]), 1)

    def test_seuls_les_brouillons_sont_exclus(self):
        """Une dépense non justifiée reste un décaissement : elle compte."""
        self.make_expense(amount="999999.00", status=Status.DRAFT)
        self.make_expense(amount="70000.00", status=Status.UNJUSTIFIED)
        self.login(self.doo)

        response = self.client.get(
            "/api/dashboard/breakdown/", {"year": self.year, "country": self.togo.pk}
        )

        total = sum(Decimal(row["amount"]) for row in response.data["by_team"])
        self.assertEqual(total, Decimal("570000.00"))

    def test_le_pays_est_obligatoire(self):
        """Sans pays, deux équipes homonymes de pays différents fusionnent."""
        Team.objects.create(country=self.ivoire, name="Équipe Lomé")
        self.login(self.doo)

        response = self.client.get("/api/dashboard/breakdown/", {"year": self.year})

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("country", response.data)

    def test_repartition_par_projet(self):
        projet = Project.objects.create(country=self.togo, name="Campagne T1")
        self.make_expense(amount="100000.00", status=Status.JUSTIFIED, project=projet)
        self.login(self.doo)

        response = self.client.get(
            "/api/dashboard/breakdown/", {"year": self.year, "country": self.togo.pk}
        )

        labels = {row["label"] for row in response.data["by_project"]}
        self.assertIn("Campagne T1", labels)
        self.assertIn("Hors projet", labels)


class AlertTests(DashboardTestCase):
    def test_seuil_franchi(self):
        # 500 000 engagés/consommés sur une enveloppe ramenée à 600 000 : 83 %.
        self.budget.amount = Decimal("600000.00")
        self.budget.save()
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        alertes = [a for a in response.data["alerts"] if a["kind"] == "budget_threshold"]
        self.assertEqual(len(alertes), 1)
        self.assertIn("80 %", alertes[0]["title"])

    def test_le_titre_suit_la_langue_demandee(self):
        """L'alerte est calculée une fois, lue dans la langue de qui la
        consulte : chaîne paresseuse, rendue à la requête."""
        self.budget.amount = Decimal("600000.00")
        self.budget.save()
        self.login(self.doo)
        # Le middleware active la langue de la requête et ne la remet pas :
        # sans ce nettoyage, les tests suivants liraient l'anglais.
        self.addCleanup(translation.deactivate)

        response = self.client.get(
            "/api/dashboard/", {"year": self.year}, HTTP_ACCEPT_LANGUAGE="en"
        )

        alerte = next(a for a in response.data["alerts"] if a["kind"] == "budget_threshold")
        self.assertEqual(alerte["title"], f"80 % threshold reached — {self.budget}")
        self.assertIn("committed or spent", alerte["detail"])

    def test_un_seul_seuil_signale_par_enveloppe(self):
        """Trois alertes pour la même enveloppe noieraient l'information."""
        self.budget.amount = Decimal("500000.00")
        self.budget.save()
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        alertes = [a for a in response.data["alerts"] if a["kind"] == "budget_threshold"]
        self.assertEqual(len(alertes), 1)
        self.assertIn("100 %", alertes[0]["title"])

    def test_depassement_signale_comme_critique(self):
        self.budget.amount = Decimal("100000.00")
        self.budget.save()
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        alerte = next(
            a for a in response.data["alerts"] if a["kind"] == "budget_overrun"
        )
        self.assertEqual(alerte["level"], "critical")

    def test_dossier_engage_sans_justificatif(self):
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        self.login(self.controller)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        alerte = next(
            a for a in response.data["alerts"] if a["kind"] == "proof_missing"
        )
        self.assertEqual(alerte["level"], "critical")
        self.assertEqual(alerte["link"], f"/dossiers/{self.dossier.pk}")

    def test_depense_inhabituelle(self):
        """Régression : incluse dans sa propre moyenne, une dépense énorme
        relevait la référence au point de ne plus s'en détacher."""
        self.make_expense(amount="50000000.00", status=Status.JUSTIFIED)
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        alerte = next(
            a for a in response.data["alerts"] if a["kind"] == "unusual_expense"
        )
        self.assertIn("50000000", alerte["title"] + alerte["detail"])

    def test_depense_ordinaire_non_signalee(self):
        self.make_expense(amount="310000.00", status=Status.JUSTIFIED)
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        self.assertFalse(
            any(a["kind"] == "unusual_expense" for a in response.data["alerts"])
        )

    def test_alertes_limitees_au_perimetre(self):
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        self.login(self.rep_ivoire)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        self.assertEqual(response.data["alerts"], [])


class NotificationTests(DashboardTestCase):
    def notifier(self):
        """Émission des alertes, telle que la planification l'exécute."""
        call_command("notify_alerts", year=self.year, verbosity=0)

    def soumettre(self, dossier=None):
        """Déclare un dossier brouillon avec une ligne, côté pays."""
        dossier = dossier or self.dossier
        expense = self.make_expense(dossier=dossier, title="Taxi", created_by=self.owner.username)
        self.login(self.owner)
        response = self.client.post(f"/api/dossiers/{dossier.pk}/submit/")
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        expense.refresh_from_db()
        return expense

    def test_consulter_le_tableau_de_bord_ne_notifie_personne(self):
        """Régression : un GET écrivait en base, et l'alerte ne partait que si
        quelqu'un ouvrait la page — personne n'était averti un dimanche."""
        self.budget.amount = Decimal("600000.00")
        self.budget.save()
        self.login(self.doo)

        response = self.client.get("/api/dashboard/", {"year": self.year})

        self.assertTrue(response.data["alerts"])
        self.assertEqual(Notification.objects.count(), 0)

    def test_seuil_budgetaire_notifie_une_seule_fois(self):
        self.budget.amount = Decimal("600000.00")
        self.budget.save()

        self.notifier()
        self.notifier()

        notifications = Notification.objects.filter(
            kind=Notification.Kind.BUDGET_THRESHOLD, recipient=self.doo
        )
        self.assertEqual(notifications.count(), 1)

    def test_justificatif_manquant_notifie_ceux_qui_peuvent_le_fournir(self):
        """Le §8 cite explicitement le justificatif manquant."""
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()

        self.notifier()

        self.assertTrue(
            Notification.objects.filter(
                recipient=self.owner, kind=Notification.Kind.PROOF_MISSING
            ).exists()
        )
        self.assertTrue(
            Notification.objects.filter(
                recipient=self.controller, kind=Notification.Kind.PROOF_MISSING
            ).exists()
        )

    def test_justificatif_manquant_notifie_une_seule_fois(self):
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()

        self.notifier()
        self.notifier()

        self.assertEqual(
            Notification.objects.filter(
                recipient=self.owner, kind=Notification.Kind.PROOF_MISSING
            ).count(),
            1,
        )

    def test_un_compte_cree_apres_coup_recoit_les_alertes_suivantes(self):
        """Un cache global de destinataires priverait d'alerte tout compte
        ouvert après le premier passage — et le second passage ne doit pas
        pour autant re-notifier ceux qui l'ont déjà été."""
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        self.notifier()
        deja = Notification.objects.filter(kind=Notification.Kind.PROOF_MISSING).count()

        nouveau = make_user("nouveau.controle", Role.DF)
        self.notifier()

        self.assertTrue(
            Notification.objects.filter(
                recipient=nouveau, kind=Notification.Kind.PROOF_MISSING
            ).exists()
        )
        self.assertEqual(
            Notification.objects.filter(kind=Notification.Kind.PROOF_MISSING).count(),
            deja + 1,
        )

    def test_soumission_du_dossier_previent_le_controleur(self):
        self.soumettre()

        notification = Notification.objects.get(
            recipient=self.controller, kind=Notification.Kind.EXPENSE_SUBMITTED
        )
        self.assertIn("N-0001", notification.title)
        # Celui qui déclare n'est pas prévenu de sa propre déclaration.
        self.assertFalse(Notification.objects.filter(recipient=self.owner).exists())

    def test_rejet_previent_le_saisisseur_avec_le_motif(self):
        # La ligne porte son auteur : c'est lui que le rejet doit prévenir, et
        # c'est aussi lui que la séparation des tâches écarte du contrôle.
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        expense = self.make_expense(
            title="Taxi", created_by=self.owner.username,
            status=Status.SUBMITTED, budget=self.budget,
        )

        self.login(self.controller)
        response = self.client.post(
            f"/api/expenses/{expense.pk}/reject/", {"note": "Reçu illisible"}
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        notification = Notification.objects.get(
            recipient=self.owner, kind=Notification.Kind.EXPENSE_REJECTED
        )
        self.assertIn("Reçu illisible", notification.body)
        self.assertEqual(notification.level, Notification.Level.WARNING)

    def test_un_email_accompagne_la_notification(self):
        self.controller.email = "dina@example.org"
        self.controller.save()

        self.soumettre()

        # Le socle de test peut donner une adresse à d'autres comptes du
        # siège : ce qui compte est que la contrôleuse reçoive le sien, seule
        # destinataire de son message.
        message = next(m for m in mail.outbox if m.to == ["dina@example.org"])
        self.assertIn("Dossier à contrôler", message.subject)
        self.assertTrue(all(len(m.to) == 1 for m in mail.outbox))

    def test_chacun_ne_voit_que_ses_notifications(self):
        """Deux destinataires réellement notifiés, chacun ne lit que sa
        ligne ; celle de l'autre n'existe pas pour lui."""
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        abidjan = Dossier.objects.create(
            number="CI-0001", label="Salon Abidjan", country=self.ivoire,
            date=date(self.year, 4, 2), status=Status.SUBMITTED,
        )
        self.notifier()  # justificatif manquant sur les deux dossiers

        self.login(self.controller)
        mine = self.client.get("/api/notifications/")
        self.login(self.rep_ivoire)
        theirs = self.client.get("/api/notifications/")
        autrui = self.client.get(f"/api/notifications/{mine.data['results'][0]['id']}/")

        self.assertEqual(mine.data["count"], 2)
        self.assertEqual(theirs.data["count"], 1)
        self.assertEqual(theirs.data["results"][0]["link"], f"/dossiers/{abidjan.pk}")
        self.assertEqual(autrui.status_code, status.HTTP_404_NOT_FOUND)

    def test_marquer_comme_lu(self):
        self.dossier.status = Status.SUBMITTED
        self.dossier.save()
        self.notifier()

        self.login(self.controller)
        before = self.client.get("/api/notifications/unread_count/")
        self.client.post("/api/notifications/read-all/")
        after = self.client.get("/api/notifications/unread_count/")

        self.assertEqual(before.data["unread"], 1)
        self.assertEqual(after.data["unread"], 0)

    def test_une_notification_ne_bloque_jamais_l_action(self):
        """Une panne d'e-mail ne doit pas empêcher de déclarer un dossier."""
        self.controller.email = "dina@example.org"
        self.controller.save()
        self.make_expense(dossier=self.dossier)
        self.login(self.owner)

        from unittest.mock import patch

        with patch(
            "notifications.services.EmailMessage.send", side_effect=OSError("SMTP HS")
        ):
            response = self.client.post(f"/api/dossiers/{self.dossier.pk}/submit/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(
            Notification.objects.filter(recipient=self.controller).exists()
        )


@in_memory_storage
class ExportTests(DashboardTestCase):
    def setUp(self):
        super().setUp()
        Proof.objects.create(
            dossier=self.dossier,
            file="justificatifs/test.pdf",
            original_name="facture.pdf",
            kind=Proof.Kind.INVOICE,
            is_complete=False,
            sha256="a" * 64,
        )

    def test_export_excel_reprend_les_colonnes_historiques(self):
        self.login(self.doo)

        response = self.client.get("/api/exports/expenses.xlsx", {"year": self.year})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        sheet = load_workbook(BytesIO(response.content)).active
        self.assertEqual(sheet.title, "BASE DE DONNEES ACTIONS")
        entetes = [cell.value for cell in sheet[1]]
        self.assertEqual(entetes[0], "N°ORDRE")
        self.assertIn("MONTANT JUSTIFIER", entetes)
        self.assertIn("PIECES JUSTIFICATIVES", entetes)

    def test_export_excel_contient_les_lignes_et_l_ecart(self):
        self.login(self.doo)

        response = self.client.get("/api/exports/expenses.xlsx", {"year": self.year})

        sheet = load_workbook(BytesIO(response.content)).active
        # Les colonnes sont retrouvées par leur en-tête : les repérer par leur
        # rang casserait au premier ajout de colonne.
        entetes = [cell.value for cell in sheet[1]]
        lignes = list(sheet.iter_rows(min_row=2, values_only=True))

        def colonne(nom):
            index = entetes.index(nom)
            return {ligne[index] for ligne in lignes}

        self.assertIn(300000.0, colonne("DEPENSES"))
        # La nuance « justif incomplet » du fichier source est conservée.
        self.assertIn(
            "Facture (justif incomplet)", colonne("PIECES JUSTIFICATIVES")
        )

    def test_un_responsable_pays_n_exporte_pas(self):
        """Le pays travaille dans l'application, sans fichier : l'export lui
        est refusé, et rien n'est tracé puisque rien n'est sorti."""
        self.login(self.rep_ivoire)

        response = self.client.get("/api/exports/expenses.xlsx", {"year": self.year})

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertFalse(AuditLog.objects.filter(object_type="Export").exists())

    def test_rapport_de_rapprochement(self):
        self.login(self.doo)

        response = self.client.get(
            "/api/exports/reconciliation.xlsx", {"year": self.year}
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames, ["Rapprochement budgets", "Rapprochement dossiers"]
        )

    def test_rapport_pdf(self):
        self.login(self.doo)

        response = self.client.get("/api/exports/report.pdf", {"year": self.year})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_export_laisse_une_trace_d_audit(self):
        """Un export sort des données du système."""
        self.login(self.doo)

        self.client.get("/api/exports/expenses.xlsx", {"year": self.year})

        entry = AuditLog.objects.filter(object_type="Export").first()
        self.assertEqual(entry.user, "do.innov")
        self.assertIn("depenses", entry.label)


class DashboardCostTests(DashboardTestCase):
    """Le coût du tableau de bord ne doit pas suivre le volume de données."""

    def _requetes(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        from core.models import WorkflowConfiguration

        # Le premier chargement de la configuration remplit le cache (six
        # requêtes) : c'est le volume que l'on mesure, pas l'amorçage.
        WorkflowConfiguration.charger()
        with CaptureQueriesContext(connection) as captured:
            response = self.client.get("/api/dashboard/", {"year": self.year})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        return len(captured.captured_queries), response

    def _peupler(self, count, offset=0):
        for index in range(offset, offset + count):
            dossier = Dossier.objects.create(
                number=f"C-{index:03d}", label=f"Dossier {index}",
                country=self.togo, date=date(self.year, 2, 1),
                status=Status.SUBMITTED,
            )
            Expense.objects.create(
                dossier=dossier, country=self.togo, budget=self.budget,
                date=f"{self.year}-02-01T10:00:00Z", title="Ligne",
                amount=Decimal("1000.00"), status=Status.JUSTIFIED,
            )

    def test_le_nombre_de_requetes_ne_suit_pas_le_volume(self):
        """Régression : chaque alerte déclenchait ses propres requêtes de
        destinataires et d'écriture — 665 requêtes pour 130 dossiers."""
        self.login(self.doo)
        self._peupler(20)
        peu, _ = self._requetes()

        self._peupler(40, offset=20)
        beaucoup, response = self._requetes()

        self.assertEqual(peu, beaucoup)
        self.assertGreater(len(response.data["alerts"]), 20)

    def test_la_charge_utile_des_alertes_est_plafonnee(self):
        from reporting.views import MAX_ALERTS

        for index in range(MAX_ALERTS + 10):
            Dossier.objects.create(
                number=f"P-{index:03d}", label=f"Dossier {index}",
                country=self.togo, date=date(self.year, 2, 1),
                status=Status.SUBMITTED,
            )
        self.login(self.doo)

        _, response = self._requetes()

        self.assertEqual(len(response.data["alerts"]), MAX_ALERTS)
        self.assertGreater(response.data["alerts_total"], MAX_ALERTS)
